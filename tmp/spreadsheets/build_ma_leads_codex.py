import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

INPUT_CSV = Path('/Users/romanduenin/Downloads/Analysis/ma_leads.csv')
OUTPUT_CSV = Path('/Users/romanduenin/Downloads/Analysis/ma_leads_codex.csv')
OUTPUT_XLSX = Path('/Users/romanduenin/Downloads/Analysis/ma_leads_codex.xlsx')
OUTPUT_SUMMARY = Path('/Users/romanduenin/Documents/New project/output/spreadsheet/ma_leads_codex_summary.txt')
ILLEGAL_XLSX_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')


def _text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and pd.isna(value):
        return ''
    return str(value).strip().lower()


def _contains_any(text: str, terms) -> bool:
    return any(term in text for term in terms)


def _clean_excel_text(value):
    if isinstance(value, str):
        return ILLEGAL_XLSX_RE.sub('', value)
    return value


def classify_row(row: pd.Series):
    industry = _text(row.get('Industry'))

    text_fields = [
        'Company Name',
        'Company Website',
        'Title',
        'Headline',
        'Industry',
        'Department',
        'Keywords',
        'Company SEO Description',
        'Company Short Description',
        'Company Domain',
    ]
    full_text = ' | '.join(_text(row.get(col)) for col in text_fields)

    yes_core_terms = [
        'private equity',
        'independent sponsor',
        'family office',
        'search fund',
        'entrepreneurship through acquisition',
        'investment bank',
        'investment banking',
        'm&a advisory',
        'm&a advisor',
        'mergers & acquisitions advisory',
        'mergers and acquisitions advisory',
        'business broker',
        'buy-side advisor',
        'sell-side advisor',
        'deal origination',
        'acquisition advisory',
        'acquisition search',
        'corporate finance advisory',
        'middle market m&a',
        'lower middle market m&a',
    ]

    ma_terms = [
        'm&a',
        'mergers & acquisitions',
        'merger and acquisition',
        'mergers and acquisitions',
        'transaction advisory',
        'deal advisory',
        'sell-side',
        'buy-side',
    ]

    acquirer_terms = [
        'we acquire',
        'actively acquires',
        'actively acquire',
        'strategic acquisitions',
        'acquisition strategy',
        'add-on acquisition',
        'platform acquisition',
        'buy-and-build',
        'buy and build',
        'looking to acquire',
        'seeking acquisitions',
        'acquisition opportunities',
        'serial acquirer',
        'acquires and operates',
    ]

    law_like = {
        'law practice',
        'legal services',
    }

    hard_no_industries = {
        'staffing & recruiting',
        'information technology & services',
        'marketing & advertising',
        'public relations & communications',
        'human resources',
        'real estate',
        'commercial real estate',
        'construction',
        'hospitality',
        'consumer services',
        'nonprofit organization management',
        'entertainment',
        'media production',
        'online media',
        'telecommunications',
        'research',
        'professional training & coaching',
    }

    finance_ambiguous = {
        'financial services',
        'investment management',
        'capital markets',
        'banking',
        'management consulting',
        'insurance',
        'accounting',
    }

    has_yes_core = _contains_any(full_text, yes_core_terms)
    has_ma_terms = _contains_any(full_text, ma_terms)
    has_acquirer_terms = _contains_any(full_text, acquirer_terms)
    has_corp_dev = 'corporate development' in full_text
    has_private_equity = 'private equity' in full_text
    has_family_office = 'family office' in full_text
    has_search_or_sponsor = (
        'search fund' in full_text
        or 'entrepreneurship through acquisition' in full_text
        or 'independent sponsor' in full_text
    )
    has_investment_banking = 'investment banking' in full_text or industry == 'investment banking'

    # Strong positives for direct ICP match.
    if has_family_office:
        return 'Yes', 'family office signal'
    if has_search_or_sponsor:
        return 'Yes', 'search fund / independent sponsor signal'
    if has_investment_banking:
        return 'Yes', 'investment banking / m&a advisor signal'

    if industry == 'venture capital & private equity':
        if has_private_equity:
            return 'Yes', 'private equity signal in vc/pe industry'
        return 'Maybe', 'vc focus but pe fit unclear'

    if industry in law_like:
        return 'No', 'legal firm, not m&a outreach buyer'

    if industry == 'accounting':
        if has_ma_terms or has_corp_dev:
            return 'Maybe', 'transaction-related accounting but advisory fit unclear'
        return 'No', 'accounting focus without m&a sourcing mandate'

    if has_yes_core:
        return 'Yes', 'explicit icp terminology present'

    if has_private_equity and industry in {'financial services', 'investment management', 'capital markets', 'banking'}:
        return 'Yes', 'private equity language in finance profile'

    if has_corp_dev or has_acquirer_terms:
        if industry in hard_no_industries:
            return 'Maybe', 'acquisition terms present but core business misaligned'
        return 'Yes', 'corporate development / acquirer language present'

    if industry == 'management consulting':
        if has_ma_terms:
            return 'Maybe', 'consulting with transaction language but not clearly advisory boutique'
        return 'No', 'general consulting without clear m&a sourcing need'

    if industry in hard_no_industries:
        return 'No', 'non-icp industry with no acquisition mandate'

    if industry in finance_ambiguous:
        if has_ma_terms:
            return 'Maybe', 'finance profile with transaction terms but unclear icp category'
        return 'Maybe', 'finance profile but no explicit icp signal'

    if has_ma_terms:
        return 'Maybe', 'm&a terms present but company type unclear'

    return 'No', 'no clear signal for deal-origination outreach need'


def apply_formatting(output_xlsx: Path, columns):
    wb = load_workbook(output_xlsx)
    ws = wb['ma_leads_codex']
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = get_column_letter(max_col)

    ws.freeze_panes = 'A2'

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sample_size = 250
    for i, col_name in enumerate(columns, start=1):
        if col_name == 'ICP':
            ws.column_dimensions[get_column_letter(i)].width = 10
            continue
        width_base = max(len(str(col_name)), 12)
        values = [ws.cell(row=r, column=i).value for r in range(2, min(max_row, sample_size) + 1)]
        max_len = max((len(str(v)) for v in values if v is not None), default=width_base)
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, max_len + 2))

    table_ref = f'A1:{end_col}{max_row}'
    tab = Table(displayName='ma_leads_codex', ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    icp_idx = list(columns).index('ICP') + 1
    icp_col = get_column_letter(icp_idx)
    data_ref = f'{icp_col}2:{icp_col}{max_row}'

    yes_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
    no_fill = PatternFill(fill_type='solid', fgColor='F8CBAD')
    maybe_fill = PatternFill(fill_type='solid', fgColor='FFE699')

    ws.conditional_formatting.add(
        data_ref,
        FormulaRule(formula=[f'${icp_col}2="Yes"'], fill=yes_fill, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        data_ref,
        FormulaRule(formula=[f'${icp_col}2="No"'], fill=no_fill, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        data_ref,
        FormulaRule(formula=[f'${icp_col}2="Maybe"'], fill=maybe_fill, stopIfTrue=True),
    )

    summary = wb.create_sheet('Summary')
    summary['A1'] = 'Metric'
    summary['B1'] = 'Value'
    summary['A2'] = 'Total rows'
    summary['B2'] = '=ROWS(ma_leads_codex[ICP])'
    summary['A3'] = 'Yes'
    summary['B3'] = '=COUNTIF(ma_leads_codex[ICP],"Yes")'
    summary['A4'] = 'No'
    summary['B4'] = '=COUNTIF(ma_leads_codex[ICP],"No")'
    summary['A5'] = 'Maybe'
    summary['B5'] = '=COUNTIF(ma_leads_codex[ICP],"Maybe")'
    summary['A7'] = 'Yes %'
    summary['B7'] = '=IF(B2=0,0,B3/B2)'
    summary['A8'] = 'No %'
    summary['B8'] = '=IF(B2=0,0,B4/B2)'
    summary['A9'] = 'Maybe %'
    summary['B9'] = '=IF(B2=0,0,B5/B2)'

    summary['D2'] = 'Most common No reason'
    summary['D3'] = 'Most common Maybe reason'

    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        cell.alignment = Alignment(horizontal='center')

    summary.column_dimensions['A'].width = 18
    summary.column_dimensions['B'].width = 14
    summary.column_dimensions['D'].width = 30
    summary.column_dimensions['E'].width = 65

    for row in ['7', '8', '9']:
        summary[f'B{row}'].number_format = '0.00%'

    wb.save(output_xlsx)


def main():
    df = pd.read_csv(INPUT_CSV)

    icp_data = df.apply(classify_row, axis=1, result_type='expand')
    icp_data.columns = ['ICP', '_ICP_Reason']

    df_out = df.copy()
    df_out['ICP'] = icp_data['ICP']
    df_xlsx = df_out.copy()
    object_cols = df_xlsx.select_dtypes(include=['object']).columns
    for col in object_cols:
        df_xlsx[col] = df_xlsx[col].map(_clean_excel_text)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    df_out.to_csv(OUTPUT_CSV, index=False)

    with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
        df_xlsx.to_excel(writer, index=False, sheet_name='ma_leads_codex')

    apply_formatting(OUTPUT_XLSX, df_out.columns)

    counts = df_out['ICP'].value_counts().to_dict()
    no_reason = icp_data.loc[icp_data['ICP'] == 'No', '_ICP_Reason'].value_counts().head(1)
    maybe_reason = icp_data.loc[icp_data['ICP'] == 'Maybe', '_ICP_Reason'].value_counts().head(1)

    wb = load_workbook(OUTPUT_XLSX)
    summary = wb['Summary']
    summary['E2'] = no_reason.index[0] if len(no_reason) else 'N/A'
    summary['E3'] = maybe_reason.index[0] if len(maybe_reason) else 'N/A'
    wb.save(OUTPUT_XLSX)

    total = len(df_out)
    yes = counts.get('Yes', 0)
    no = counts.get('No', 0)
    maybe = counts.get('Maybe', 0)

    report = [
        f'Total: {total}',
        f'Yes: {yes}',
        f'No: {no}',
        f'Maybe: {maybe}',
        f"Most common No reason: {summary['E2'].value}",
        f"Most common Maybe reason: {summary['E3'].value}",
        f'CSV: {OUTPUT_CSV}',
        f'XLSX: {OUTPUT_XLSX}',
    ]
    OUTPUT_SUMMARY.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SUMMARY.write_text('\n'.join(report), encoding='utf-8')

    print('\n'.join(report))


if __name__ == '__main__':
    main()
